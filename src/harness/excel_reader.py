from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class OptimizationRecord:
    scene_raw: str
    scene_key: str
    content: str
    sub_scene_raw: str
    sub_scene_key: str
    domain: str
    row_index: int
    source: str = "xlsx"


@dataclass(frozen=True)
class OptimizationGroupKey:
    scene_key: str
    sub_scene_key: str = ""


def read_optimization_records(path: str | Path) -> list[OptimizationRecord]:
    data_path = Path(path)
    if data_path.suffix.lower() == ".json":
        return read_optimization_records_json(data_path)
    return read_optimization_records_xlsx(data_path)


def read_optimization_records_xlsx(path: str | Path) -> list[OptimizationRecord]:
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Optimization data file not found: {xlsx_path}")
    if xlsx_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx and .json optimization data is supported: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    records: list[OptimizationRecord] = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        scene_raw = _cell_text(row[0] if len(row) > 0 else "")
        content = _cell_text(row[2] if len(row) > 2 else "")
        sub_scene_raw = _cell_text(row[20] if len(row) > 20 else "")
        domain = _cell_text(row[21] if len(row) > 21 else "")
        if row_index == 1 and _looks_like_header(scene_raw, content, sub_scene_raw, domain):
            continue
        if not scene_raw and not content and not sub_scene_raw and not domain:
            continue
        if not scene_raw:
            scene_raw = "unknown"
        records.append(
            OptimizationRecord(
                scene_raw=scene_raw,
                scene_key=normalize_scene_key(scene_raw),
                content=content,
                sub_scene_raw=sub_scene_raw,
                sub_scene_key=normalize_sub_scene_key(sub_scene_raw),
                domain=domain,
                row_index=row_index,
                source="xlsx",
            )
        )

    workbook.close()
    return records


def read_optimization_records_json(path: str | Path) -> list[OptimizationRecord]:
    json_path = Path(path)
    if not json_path.exists():
        raise FileNotFoundError(f"Optimization data file not found: {json_path}")
    if json_path.suffix.lower() != ".json":
        raise ValueError(f"Only .json optimization data is supported: {json_path}")

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    records: list[OptimizationRecord] = []
    for index, item in enumerate(_extract_json_items(payload), start=1):
        scene_raw = _json_text(item, "scene", "场景", "scene_raw")
        content = _json_text(item, "content", "发言内容", "utterance", "speech", "text")
        sub_scene_raw = _json_text(item, "sub_scene", "subScene", "subscene", "子场景")
        domain = _json_text(item, "domain", "领域")
        if not scene_raw and not content and not sub_scene_raw and not domain:
            continue
        if not scene_raw:
            scene_raw = "unknown"
        records.append(
            OptimizationRecord(
                scene_raw=scene_raw,
                scene_key=normalize_scene_key(scene_raw),
                content=content,
                sub_scene_raw=sub_scene_raw,
                sub_scene_key=normalize_sub_scene_key(sub_scene_raw),
                domain=domain,
                row_index=index,
                source="json",
            )
        )
    return records


def group_records_by_scene(records: list[OptimizationRecord]) -> dict[str, list[OptimizationRecord]]:
    grouped: dict[str, list[OptimizationRecord]] = {}
    for record in records:
        grouped.setdefault(record.scene_key, []).append(record)
    return grouped


def group_records_by_scene_and_sub_scene(
    records: list[OptimizationRecord],
) -> dict[OptimizationGroupKey, list[OptimizationRecord]]:
    return group_records_by_optimization_scope(records=records, scope="scene_sub_scene")


def group_records_by_optimization_scope(
    records: list[OptimizationRecord],
    scope: str,
) -> dict[OptimizationGroupKey, list[OptimizationRecord]]:
    valid_scopes = {"scene", "sub_scene", "scene_sub_scene", "scene_and_sub_scene"}
    if scope not in valid_scopes:
        raise ValueError("Optimization scope must be one of: scene, sub_scene, scene_sub_scene, scene_and_sub_scene.")
    grouped: dict[OptimizationGroupKey, list[OptimizationRecord]] = {}
    for record in records:
        if scope == "scene":
            key = OptimizationGroupKey(scene_key=record.scene_key)
        elif scope == "sub_scene":
            key = OptimizationGroupKey(scene_key="", sub_scene_key=record.sub_scene_key or "unknown")
        elif scope == "scene_and_sub_scene":
            grouped.setdefault(OptimizationGroupKey(scene_key=record.scene_key), []).append(record)
            if record.sub_scene_key:
                key = OptimizationGroupKey(scene_key=record.scene_key, sub_scene_key=record.sub_scene_key)
            else:
                continue
        else:
            key = OptimizationGroupKey(scene_key=record.scene_key, sub_scene_key=record.sub_scene_key)
        grouped.setdefault(key, []).append(record)
    return grouped


def build_scene_context(
    scene_key: str,
    records: list[OptimizationRecord],
    sub_scene_key: str = "",
) -> str:
    lines = [f"# Scene: {scene_key}"]
    if sub_scene_key:
        lines.append(f"Sub-scene: {sub_scene_key}")
    domains = sorted({record.domain for record in records if record.domain})
    if domains:
        lines.append(f"Domains: {', '.join(domains)}")
    lines.extend([f"Sample count: {len(records)}", ""])
    for idx, record in enumerate(records, start=1):
        lines.append(f"## Sample {idx}")
        lines.append(f"Raw scene: {record.scene_raw}")
        lines.append(f"Sub-scene: {record.sub_scene_raw or '(empty sub-scene)'}")
        lines.append(f"Domain: {record.domain or '(empty domain)'}")
        lines.append("Content:")
        lines.append(record.content or "(empty content)")
        lines.append("")
    return "\n".join(lines).strip()


def normalize_scene_key(scene_raw: str) -> str:
    text = scene_raw.strip()
    text = re.sub(r"[\s_-]*[0-9０-９]+$", "", text)
    text = re.sub(r"[\s_-]+$", "", text)
    return text or scene_raw.strip() or "unknown"


def normalize_sub_scene_key(sub_scene_raw: str) -> str:
    if not sub_scene_raw.strip():
        return ""
    return normalize_scene_key(sub_scene_raw)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_header(scene: str, content: str, sub_scene: str, domain: str) -> bool:
    scene_l = scene.lower().strip()
    content_l = content.lower().strip()
    sub_scene_l = sub_scene.lower().strip()
    domain_l = domain.lower().strip()
    scene_tokens = {"scene", "type", "scene type", "场景", "类型", "场景类型"}
    content_tokens = {"content", "text", "utterance", "speech", "发言内容", "内容", "正文"}
    sub_scene_tokens = {"sub_scene", "subscene", "sub scene", "子场景", "子类"}
    domain_tokens = {"domain", "field", "领域"}
    return (
        scene_l in scene_tokens
        and content_l in content_tokens
        and (not sub_scene_l or sub_scene_l in sub_scene_tokens)
        and (not domain_l or domain_l in domain_tokens)
    )


def _extract_json_items(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        raise ValueError("JSON optimization data must contain an object or an array of objects.")
    if _has_record_fields(payload):
        return [payload]
    for key in ("records", "items", "data", "samples"):
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    items: list[dict[str, Any]] = []
    for scene, value in payload.items():
        if isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    items.append({"scene": scene, **item})
        elif isinstance(value, dict):
            for sub_scene, nested in value.items():
                if isinstance(nested, list):
                    for item in nested:
                        if isinstance(item, dict):
                            items.append({"scene": scene, "sub_scene": sub_scene, **item})
    if items:
        return items
    raise ValueError("JSON optimization data does not contain readable records.")


def _has_record_fields(item: dict[str, Any]) -> bool:
    field_names = {
        "scene",
        "场景",
        "content",
        "发言内容",
        "utterance",
        "speech",
        "text",
        "sub_scene",
        "subScene",
        "subscene",
        "子场景",
        "domain",
        "领域",
    }
    return any(key in item for key in field_names)


def _json_text(item: dict[str, Any], *keys: str) -> str:
    for key in keys:
        if key in item:
            return _cell_text(item[key])
    return ""
